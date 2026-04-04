"""
XLSX 格式水印处理：嵌入/提取零宽字符。

仅负责文件格式操作（读写 XLSX 单元格文本），
编解码逻辑由 office_wm.py 统一处理。
"""

from pathlib import Path
from typing import Optional

import openpyxl
from loguru import logger


def embed_xlsx(
    input_path: Path, zwc_block: str, output_path: Path,
) -> tuple[bool, str]:
    """
    在 XLSX 文件中嵌入零宽字符块。

    找到第一个非空字符串单元格，在其值前端插入 ZWC 块。

    Returns:
        (success, message) 元组
    """
    try:
        wb = openpyxl.load_workbook(str(input_path))
    except Exception as e:
        return False, f"Cannot open XLSX: {e}"

    inserted = False
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip():
                    cell.value = zwc_block + cell.value
                    inserted = True
                    break
            if inserted:
                break
        if inserted:
            break

    if not inserted:
        wb.close()
        return False, "No text content found in XLSX"

    try:
        wb.save(str(output_path))
    except Exception as e:
        wb.close()
        return False, f"Cannot save XLSX: {e}"

    wb.close()
    logger.debug(f"ZWC block inserted into XLSX: {input_path.name}")
    return True, "OK"


def extract_xlsx(file_path: Path) -> Optional[str]:
    """
    从 XLSX 文件提取所有字符串单元格内容。

    扫描所有 sheet 的所有字符串 cell，拼接文本供 zwc_decode 搜索。

    Returns:
        拼接后的全部字符串内容，失败返回 None
    """
    try:
        wb = openpyxl.load_workbook(str(file_path), read_only=True)
    except Exception as e:
        logger.warning(f"Cannot open XLSX for extraction: {e}")
        return None

    parts = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    parts.append(cell.value)
    wb.close()
    return "".join(parts) if parts else None
